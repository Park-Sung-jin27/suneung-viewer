# -*- coding: utf-8 -*-
"""
adiga 5년치 raw → 학과별 5년 추이 추출
입시데이터_마스터_v2.0.xlsx (adiga_syr2022~2026_raw) → js/trend_data.js

기존 가천대 정형 추출 + adiga 5년치 추출 통합.
"""
import openpyxl
import json
import re
from pathlib import Path
from collections import defaultdict

SRC = Path(__file__).parent.parent / "입시데이터_마스터_v2.0.xlsx"
OUT = Path(__file__).parent / "js" / "trend_data.js"

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)


def parse_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


# ===== 가천대 정형 (기존 로직 재사용) =====
def extract_gachon():
    ws = wb["가천대_다년치_원본"]
    by_dept = defaultdict(list)
    for r in ws.iter_rows(values_only=True, min_row=2):
        if not r:
            continue
        year = r[1]
        type_ = r[2]
        cells = r[5:]
        if not cells or not cells[2]:
            continue
        dept = cells[2]
        if not isinstance(dept, str) or len(dept) < 2:
            continue
        if "(야)" in dept or "야간" in dept:
            continue

        if type_ == "정시":
            trans_name = cells[0] if isinstance(cells[0], str) else ""
            capacity = cells[4] if isinstance(cells[4], (int, float)) else None
            comp = cells[6] if isinstance(cells[6], (int, float)) else None
            cut = cells[7] if isinstance(cells[7], (int, float)) else None
            if cut is not None and (cut < 30 or cut > 100):
                cut = None
        elif type_ == "수시":
            trans_name = cells[1] if isinstance(cells[1], str) else ""
            capacity = cells[3] if isinstance(cells[3], (int, float)) else None
            comp = cells[4] if isinstance(cells[4], (int, float)) else None
            cut = cells[5] if isinstance(cells[5], (int, float)) else None
            if cut is not None and (cut < 1 or cut > 9):
                cut = None
            if "논술" in trans_name:
                continue
        else:
            continue
        if any(kw in trans_name for kw in ["농어촌", "기회균형", "교육기회", "특수교육", "재외국민", "특성화고"]):
            continue
        if capacity is None or comp is None or cut is None:
            continue
        by_dept[(dept, type_)].append({
            "year": year, "전형": trans_name, "capacity": capacity,
            "competition": round(comp, 2), "cut70": round(cut, 2),
        })

    aggregated = {}
    for (dept, type_), recs in by_dept.items():
        if len(recs) < 2:
            continue
        ygroup = defaultdict(list)
        for r in recs:
            ygroup[r["year"]].append(r)
        agg = []
        for year, items in sorted(ygroup.items()):
            if year is None:
                continue
            cuts = [x["cut70"] for x in items if x["cut70"] is not None]
            comps = [x["competition"] for x in items if x["competition"] is not None]
            caps = [x["capacity"] for x in items if x["capacity"] is not None]
            if not cuts:
                continue
            agg.append({
                "year": year, "cut70": round(sum(cuts) / len(cuts), 2),
                "competition": round(sum(comps) / len(comps), 2) if comps else None,
                "capacity": sum(caps),
            })
        if len(agg) >= 2:
            key = "가천대학교|" + dept + "|" + type_
            aggregated[key] = agg
    return aggregated


# ===== adiga 5년치 추출 =====
# searchSyr → 학년도 (학년도 = searchSyr - 1)
SHEETS = {
    "adiga_syr2022_raw": 2021,
    "adiga_syr2023_raw": 2022,
    "adiga_syr2024_raw": 2023,
    "adiga_syr2025_raw": 2024,
    "adiga_syr2026_raw": 2025,
}

NOISE_DEPTS = ("총계", "소계", "합계", "% cut", "50% cut", "70% cut", "90% cut")
SPECIAL_TRANS = ("농어촌", "기회균형", "기회 균형", "교육기회", "특수교육", "재외국민",
                 "특성화고", "북한", "탈북", "다문화", "사회배려", "장애", "보훈")


def extract_adiga_sheet(sheet_name, year_value):
    ws = wb[sheet_name]
    # 표별 그룹: (대학, 캠퍼스, table_idx)
    tables = defaultdict(list)
    for r in ws.iter_rows(values_only=True, min_row=2):
        if not r:
            continue
        univ, campus, syr, table_idx, row_idx = r[0], r[1], r[2], r[3], r[4]
        cells = r[5:]
        if univ is None or table_idx is None:
            continue
        tables[(univ, campus, table_idx)].append((row_idx, cells))

    out = []
    for (univ, campus, table_idx), rows in tables.items():
        rows.sort(key=lambda x: x[0])
        # row_idx 0 = 전형명 (c2 in cells = idx 1)
        # row_idx 1 = 컬럼 헤더
        # row_idx >= 2 = 데이터
        trans_name = ""
        header = []
        data_rows = []
        for ridx, cells in rows:
            if ridx == 0 and len(cells) > 1 and isinstance(cells[1], str):
                trans_name = cells[1].strip()
            elif ridx == 1:
                header = list(cells)
            elif ridx >= 2:
                data_rows.append(cells)
        if not header or not data_rows:
            continue
        # 특별 전형 제외
        if any(k in trans_name for k in SPECIAL_TRANS):
            continue
        # cut 컬럼 위치 찾기 — "70% cut" 또는 "70%cut" 또는 "교과성적" 포함된 헤더
        cut_col = None
        for i, h in enumerate(header):
            if h is None:
                continue
            hs = str(h).replace(" ", "")
            if "70%cut" in hs or "70cut" in hs:
                cut_col = i
                break
        if cut_col is None:
            # 단일 cut 컬럼: "최종등록자" 포함 + 교과성적/학생부등급
            for i, h in enumerate(header):
                if h is None:
                    continue
                hs = str(h).replace(" ", "")
                if ("최종등록자" in hs and ("교과성적" in hs or "학생부등급" in hs or "환산등급" in hs or "cut" in hs)) or hs == "cut":
                    cut_col = i
                    break
        if cut_col is None:
            continue
        # header에서 모집인원·경쟁률 컬럼 (보통 c2=모집인원=헤더idx 0, c3=경쟁률=헤더idx 1, c4=충원=idx 2, c5=cut=idx 3)
        # 데이터 row: cells[0]=학과명, cells[1]=모집인원, cells[2]=경쟁률, cells[cut_col+1?]=cut70
        # 실제 데이터 row는 컬럼이 헤더와 같은 idx에 옴 — 그러나 학과명이 cells[0]에 추가됨
        # → 학과명 = cells[0], 모집인원 = cells[1], 경쟁률 = cells[2], cut = cells[cut_col + 1]
        cap_idx = 1
        comp_idx = 2
        cut_idx = cut_col + 1

        for data in data_rows:
            if not data or not data[0]:
                continue
            dept = str(data[0]).strip()
            if any(n in dept for n in NOISE_DEPTS):
                continue
            if len(dept) < 2:
                continue
            cap = parse_num(data[cap_idx]) if cap_idx < len(data) else None
            comp = parse_num(data[comp_idx]) if comp_idx < len(data) else None
            cut = parse_num(data[cut_idx]) if cut_idx < len(data) else None
            if cap is None or cut is None:
                continue
            out.append({
                "univ": univ, "campus": campus or "본", "dept": dept,
                "전형": trans_name, "year": year_value,
                "capacity": cap, "competition": comp, "cut70": cut,
            })
    return out


def normalize_univ(name):
    if not name:
        return name
    if name.endswith("여대"):
        return name.replace("여대", "여자대학교")
    if name == "한국외대":
        return "한국외국어대학교"
    if name.endswith("대"):
        return name[:-1] + "대학교"
    return name


def aggregate_adiga():
    # 5년치 모든 record
    all_records = []
    for sheet, year in SHEETS.items():
        recs = extract_adiga_sheet(sheet, year)
        all_records.extend(recs)
        print("  " + sheet + ": " + str(len(recs)) + " records")

    # 그룹화 — (대학, 학과) 키로
    grouped = defaultdict(list)
    for r in all_records:
        univ_norm = normalize_univ(r["univ"])
        # cut 단위 추정: 1~9 = 등급(수시), 30~100 = 환산점수%
        # 같은 (대학,학과)에 등급과 점수 둘 다 있을 수 있음 — 구분
        cut = r["cut70"]
        if 1 <= cut <= 9:
            type_ = "수시"
        elif 30 <= cut <= 1500:
            type_ = "정시"
        else:
            continue
        key = (univ_norm, r["dept"], type_)
        grouped[key].append(r)

    aggregated = {}
    for (univ, dept, type_), recs in grouped.items():
        if len(recs) < 2:
            continue
        # year별 평균
        ygroup = defaultdict(list)
        for r in recs:
            ygroup[r["year"]].append(r)
        agg = []
        for year, items in sorted(ygroup.items()):
            cuts = [x["cut70"] for x in items if x["cut70"] is not None]
            comps = [x["competition"] for x in items if x["competition"] is not None]
            caps = [x["capacity"] for x in items if x["capacity"] is not None]
            if not cuts:
                continue
            agg.append({
                "year": year, "cut70": round(sum(cuts) / len(cuts), 2),
                "competition": round(sum(comps) / len(comps), 2) if comps else None,
                "capacity": sum(caps),
            })
        if len(agg) >= 2:
            key = univ + "|" + dept + "|" + type_
            aggregated[key] = agg
    return aggregated


def main():
    trend = {}

    g = extract_gachon()
    print("가천대 정형: " + str(len(g)))
    trend.update(g)

    print("adiga 5년치 추출:")
    a = aggregate_adiga()
    print("adiga 추이 학과: " + str(len(a)))

    # 가천대는 정형 우선 (덮어쓰기 X)
    for k, v in a.items():
        if k not in trend:
            trend[k] = v

    by_univ = defaultdict(int)
    by_type = defaultdict(int)
    by_yc = defaultdict(int)
    for k, v in trend.items():
        univ, _, t = k.split("|")
        by_univ[univ] += 1
        by_type[t] += 1
        by_yc[len(v)] += 1
    print("총 학과: " + str(len(trend)))
    print("상위 대학: " + str(sorted(by_univ.items(), key=lambda x: -x[1])[:10]))
    print("구분별:", dict(by_type))
    print("연수별:", dict(sorted(by_yc.items())))

    js = ("// 자동 생성 — extract_trend.py 출력\n"
          "window.TREND_DATA = " + json.dumps(trend, ensure_ascii=False, indent=1) + ";\n")
    OUT.write_text(js, encoding="utf-8")
    print("trend_data.js: " + str(len(js)) + " bytes")


if __name__ == "__main__":
    main()
