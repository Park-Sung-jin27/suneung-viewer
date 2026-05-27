# -*- coding: utf-8 -*-
"""
extract_data.py — 데모 데이터 추출 (v2)
입시데이터_마스터_v2.0.xlsx + adiga raw 5년치 통합 → js/data.js

추가:
  - 정시 컷 누락 대학을 adiga raw에서 최신(2025) 컷으로 보강
  - 수시 컷 누락 대학도 동일
"""
import openpyxl
import json
from pathlib import Path
from collections import defaultdict

SRC = Path(__file__).parent.parent / "입시데이터_마스터_v2.0.xlsx"
OUT_DIR = Path(__file__).parent / "js"
OUT_DIR.mkdir(exist_ok=True)

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)


def rows(sheet_name, header_row_idx):
    ws = wb[sheet_name]
    data = list(ws.iter_rows(values_only=True))
    headers = data[header_row_idx]
    out = []
    for r in data[header_row_idx + 1:]:
        if r is None or all(v is None for v in r):
            continue
        out.append(dict(zip(headers, r)))
    return out


def normalize_univ(name):
    if not name:
        return name
    if name.endswith("여대"):
        return name.replace("여대", "여자대학교")
    if name == "한국외대":
        return "한국외국어대학교"
    if name.endswith("대"):
        return name[:-1] + "대학교"
    return name


def parse_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


# ===== universities =====
unis_raw = rows("대학_마스터", 0)
universities = []
for u in unis_raw:
    if not u.get("대학명"):
        continue
    universities.append({
        "name": u["대학명"],
        "campus": u.get("본교_분교") or "본교",
        "unvCd": u.get("adiga_unvCd"),
        "region": u.get("지역"),
        "type": u.get("설립유형"),
        "competition_susi_25": u.get("2025_수시경쟁률"),
        "competition_jeongsi_25": u.get("2025_정시경쟁률"),
        "intake": u.get("입학정원"),
    })
print(f"universities: {len(universities)}")


# ===== conversion_formulas =====
formulas_raw = rows("환산공식_정시", 0)
formulas = []
for f in formulas_raw:
    if not f.get("대학명"):
        continue
    formulas.append({
        "univ": f["대학명"], "campus": f.get("캠퍼스") or "본교",
        "track": f.get("계열"),
        "weights": {"국어": f.get("국어_반영%") or 0, "수학": f.get("수학_반영%") or 0,
                    "영어": f.get("영어_반영%") or 0, "탐구": f.get("탐구_반영%") or 0,
                    "한국사": f.get("한국사_반영%") or 0},
        "metric": {"국어": f.get("국어_지표"), "수학": f.get("수학_지표"),
                   "탐구": f.get("탐구_지표")},
        "english_grade_table": [f.get(f"영어_{i}등급") for i in range(1, 10)],
        "korean_history_grade_table": [f.get(f"한국사_{i}등급") for i in range(1, 10)],
        "math_constraint": f.get("수학_선택지정"),
        "explore_constraint": f.get("탐구_영역지정"),
        "bonus": f.get("가산점"),
        "max_score": f.get("만점"),
        "source": f.get("출처_검증필요") or "잠정",
    })
print(f"formulas: {len(formulas)}")


# ===== jeongsi_cutoffs (기존 시트) =====
js_raw = rows("학과별_정시_합격가능성", 1)
jeongsi_cutoffs = []
existing_jeongsi_keys = set()
for r in js_raw:
    if not r.get("대학명") or not r.get("학과명"):
        continue
    cut = r.get("작년_70cut_환산점수")
    if cut is None or cut == "":
        continue
    jeongsi_cutoffs.append({
        "univ": r["대학명"], "campus": r.get("캠퍼스") or "본교",
        "dept": r["학과명"], "group": r.get("모집군"),
        "capacity": r.get("모집인원"), "competition": r.get("경쟁률"),
        "cut70": cut, "source": r.get("source"),
    })
    existing_jeongsi_keys.add((normalize_univ(r["대학명"]), r["학과명"]))
print(f"jeongsi_cutoffs (기존): {len(jeongsi_cutoffs)}")


# ===== adiga raw 5년치 → 누락 대학 정시 보강 =====
SHEETS = {
    "adiga_syr2022_raw": 2021, "adiga_syr2023_raw": 2022, "adiga_syr2024_raw": 2023,
    "adiga_syr2025_raw": 2024, "adiga_syr2026_raw": 2025,
}

NOISE_DEPTS = ("총계", "소계", "합계", "% cut", "50% cut", "70% cut", "90% cut")
SPECIAL_TRANS = ("농어촌", "기회균형", "기회 균형", "교육기회", "특수교육", "재외국민",
                 "특성화고", "북한", "탈북", "다문화", "사회배려", "장애", "보훈")


def extract_adiga_sheet(sheet_name, year_value):
    ws = wb[sheet_name]
    tables = defaultdict(list)
    for r in ws.iter_rows(values_only=True, min_row=2):
        if not r:
            continue
        univ, campus, syr, table_idx, row_idx = r[0], r[1], r[2], r[3], r[4]
        cells = r[5:]
        if univ is None or table_idx is None:
            continue
        tables[(univ, campus, table_idx)].append((row_idx, cells))

    out = []
    for (univ, campus, table_idx), trows in tables.items():
        trows.sort(key=lambda x: x[0])
        trans_name = ""
        header = []
        data_rows = []
        for ridx, cells in trows:
            if ridx == 0 and len(cells) > 1 and isinstance(cells[1], str):
                trans_name = cells[1].strip()
            elif ridx == 1:
                header = list(cells)
            elif ridx >= 2:
                data_rows.append(cells)
        if not header or not data_rows:
            continue
        if any(k in trans_name for k in SPECIAL_TRANS):
            continue

        cut_col = None
        for i, h in enumerate(header):
            if h is None:
                continue
            hs = str(h).replace(" ", "")
            if "70%cut" in hs or "70cut" in hs:
                cut_col = i
                break
        if cut_col is None:
            for i, h in enumerate(header):
                if h is None:
                    continue
                hs = str(h).replace(" ", "")
                if ("최종등록자" in hs and ("교과성적" in hs or "학생부등급" in hs or "환산등급" in hs)) or hs == "cut":
                    cut_col = i
                    break
        if cut_col is None:
            continue

        cap_idx = 1
        comp_idx = 2
        cut_idx = cut_col + 1
        for data in data_rows:
            if not data or not data[0]:
                continue
            dept = str(data[0]).strip()
            if any(n in dept for n in NOISE_DEPTS):
                continue
            if len(dept) < 2:
                continue
            cap = parse_num(data[cap_idx]) if cap_idx < len(data) else None
            comp = parse_num(data[comp_idx]) if comp_idx < len(data) else None
            cut = parse_num(data[cut_idx]) if cut_idx < len(data) else None
            if cap is None or cut is None:
                continue
            out.append({
                "univ": univ, "campus": campus or "본", "dept": dept,
                "전형": trans_name, "year": year_value,
                "capacity": cap, "competition": comp, "cut70": cut,
            })
    return out


# 모든 sheet에서 record 모음 (정시·수시 모두)
all_adiga = []
for sh, yr in SHEETS.items():
    recs = extract_adiga_sheet(sh, yr)
    all_adiga.extend(recs)
print(f"adiga 전체 record: {len(all_adiga)}")

# 최신 학년도 (각 대학·학과별 가장 최근) 우선
best_by_key = {}
for r in all_adiga:
    cut = r["cut70"]
    if 1 <= cut <= 9:
        type_ = "수시"
    elif 30 <= cut <= 1500:
        type_ = "정시"
    else:
        continue
    key = (normalize_univ(r["univ"]), r["dept"], type_)
    if key not in best_by_key or r["year"] > best_by_key[key]["year"]:
        best_by_key[key] = {**r, "type": type_}

# 정시 추가 (기존에 없는 (대학, 학과) 만)
added_jeongsi = 0
for (uniN, dept, type_), rec in best_by_key.items():
    if type_ != "정시":
        continue
    if (uniN, dept) in existing_jeongsi_keys:
        continue
    jeongsi_cutoffs.append({
        "univ": uniN, "campus": "본교", "dept": dept,
        "group": "정시", "capacity": rec["capacity"],
        "competition": rec.get("competition"),
        "cut70": rec["cut70"],
        "source": "adiga_" + str(rec["year"]),
    })
    added_jeongsi += 1
print(f"jeongsi_cutoffs (보강 후): {len(jeongsi_cutoffs)} (+{added_jeongsi})")


# ===== susi_cutoffs (기존 + 보강) =====
ss_raw = rows("학과별_수시_합격가능성", 1)
susi_cutoffs = []
existing_susi_keys = set()
for r in ss_raw:
    if not r.get("대학명") or not r.get("학과명"):
        continue
    cut = r.get("작년_70cut")
    if cut is None or cut == "":
        continue
    susi_cutoffs.append({
        "univ": r["대학명"], "campus": r.get("캠퍼스") or "본교",
        "dept": r["학과명"], "전형": r.get("전형명"),
        "capacity": r.get("모집인원"), "competition": r.get("경쟁률"),
        "cut70_grade": cut, "source": r.get("source"),
    })
    existing_susi_keys.add((normalize_univ(r["대학명"]), r["학과명"], r.get("전형명") or ""))

added_susi = 0
for (uniN, dept, type_), rec in best_by_key.items():
    if type_ != "수시":
        continue
    if (uniN, dept, rec.get("전형") or "") in existing_susi_keys:
        continue
    susi_cutoffs.append({
        "univ": uniN, "campus": "본교", "dept": dept,
        "전형": rec.get("전형"),
        "capacity": rec["capacity"], "competition": rec.get("competition"),
        "cut70_grade": rec["cut70"], "source": "adiga_" + str(rec["year"]),
    })
    added_susi += 1
print(f"susi_cutoffs (보강 후): {len(susi_cutoffs)} (+{added_susi})")


# ===== sample_students =====
sample_students = [
    {"id": "S001", "name": "김민수", "grade": "고3", "track": "자연",
     "scores": {"국어": {"raw": 88, "std": 130, "percentile": 92, "grade": 2},
                "수학": {"elective": "미적분", "raw": 92, "std": 135, "percentile": 95, "grade": 2},
                "영어": {"raw": 85, "grade": 2},
                "탐구1": {"subject": "물리1", "raw": 42, "std": 65, "percentile": 88, "grade": 3},
                "탐구2": {"subject": "화학1", "raw": 45, "std": 67, "percentile": 90, "grade": 2},
                "한국사": {"raw": 35, "grade": 3}},
     "내신": 2.8, "수상": ["과학탐구대회 은상", "수학경시 동상"],
     "희망_계열": ["공학", "자연과학"], "희망_지역": ["서울", "수도권"],
     "제약": ["통학 1시간 이내", "등록금 1000만원 이하"],
     "이탈위험": False, "마지막상담": "2026-05-15"},
    {"id": "S002", "name": "이지은", "grade": "고3", "track": "인문",
     "scores": {"국어": {"raw": 93, "std": 138, "percentile": 96, "grade": 1},
                "수학": {"elective": "확률과통계", "raw": 78, "std": 118, "percentile": 80, "grade": 3},
                "영어": {"raw": 92, "grade": 1},
                "탐구1": {"subject": "생활과윤리", "raw": 47, "std": 68, "percentile": 93, "grade": 2},
                "탐구2": {"subject": "사회문화", "raw": 45, "std": 66, "percentile": 91, "grade": 2},
                "한국사": {"raw": 42, "grade": 2}},
     "내신": 2.3, "수상": ["토론대회 금상", "독서감상문 우수상"],
     "희망_계열": ["경영", "경제", "심리"], "희망_지역": ["서울"],
     "제약": ["통학 30분 이내"], "이탈위험": True, "마지막상담": "2026-04-28"},
    {"id": "S004", "name": "최예진", "grade": "고3", "track": "자연",
     "scores": {"국어": {"raw": 75, "std": 116, "percentile": 75, "grade": 4},
                "수학": {"elective": "미적분", "raw": 78, "std": 120, "percentile": 78, "grade": 4},
                "영어": {"raw": 80, "grade": 3},
                "탐구1": {"subject": "생명1", "raw": 38, "std": 62, "percentile": 75, "grade": 4},
                "탐구2": {"subject": "화학1", "raw": 40, "std": 63, "percentile": 78, "grade": 4},
                "한국사": {"raw": 35, "grade": 4}},
     "내신": 3.5, "수상": ["SW 경진대회 동상"],
     "희망_계열": ["IT", "공학"], "희망_지역": ["수도권"],
     "제약": ["통학 1시간 이내"], "이탈위험": False, "마지막상담": "2026-05-20"},
    {"id": "S003", "name": "박서준", "grade": "고3", "track": "자연",
     "scores": {"국어": {"raw": 75, "std": 115, "percentile": 75, "grade": 4},
                "수학": {"elective": "미적분", "raw": 80, "std": 122, "percentile": 82, "grade": 3},
                "영어": {"raw": 78, "grade": 3},
                "탐구1": {"subject": "생명1", "raw": 38, "std": 62, "percentile": 78, "grade": 4},
                "탐구2": {"subject": "지구과학1", "raw": 40, "std": 64, "percentile": 82, "grade": 3},
                "한국사": {"raw": 38, "grade": 3}},
     "내신": 3.4, "수상": [],
     "희망_계열": ["IT", "컴퓨터공학"], "희망_지역": ["수도권"],
     "제약": [], "이탈위험": False, "마지막상담": "2026-05-18"},
]
print(f"students: {len(sample_students)}")


data_bundle = {
    "universities": universities, "formulas": formulas,
    "jeongsi_cutoffs": jeongsi_cutoffs, "susi_cutoffs": susi_cutoffs,
    "students": sample_students,
    "meta": {"extracted_from": "v2.0 + adiga raw", "extracted_at": "2026-05-26",
             "rule_version": "2027.5",
             "disclaimer": "환산공식 30개대학 잠정. 정시·수시 컷은 adiga raw 5년치 통합 (최신 학년도 우선)."}
}

js_content = ("// 자동 생성 — extract_data.py 출력\n"
              "window.DEMO_DATA = " + json.dumps(data_bundle, ensure_ascii=False, indent=1) + ";\n")
(OUT_DIR / "data.js").write_text(js_content, encoding="utf-8")
print(f"data.js: {len(js_content):,} bytes")
